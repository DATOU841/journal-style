#!/usr/bin/env python3
"""Analyze public author and institution networks from title metadata."""

from __future__ import annotations

import argparse
import csv
import itertools
import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path


SPLIT_PATTERN = re.compile(r"[;；,，、/|]+")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Analyze author and institution network from metadata.")
    parser.add_argument("--input", required=True, help="CSV/JSON/XLSX title metadata.")
    parser.add_argument("--output-json", default="journal-author-institution-network-statistics.json")
    parser.add_argument("--output-md", default="journal-author-institution-network-report.md")
    return parser.parse_args()


def load_records(path: Path) -> list[dict]:
    suffix = path.suffix.lower()
    if suffix == ".json":
        data = json.loads(path.read_text(encoding="utf-8"))
        return list(data if isinstance(data, list) else data.get("records", []))
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as fh:
            return list(csv.DictReader(fh))
    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except Exception as exc:  # pragma: no cover - dependency guard
            raise SystemExit("openpyxl is required for xlsx input") from exc
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        headers = [str(cell).strip() if cell is not None else "" for cell in next(rows)]
        return [{headers[i]: row[i] for i in range(min(len(headers), len(row)))} for row in rows]
    raise SystemExit(f"Unsupported input format: {path.suffix}")


def pick(row: dict, *keys: str) -> str:
    lowered = {str(k).strip().lower(): v for k, v in row.items()}
    for key in keys:
        value = lowered.get(key.lower())
        if value is not None and str(value).strip():
            return str(value).strip()
    return ""


def split_people(value: str) -> list[str]:
    items = []
    for item in SPLIT_PATTERN.split(value or ""):
        clean = re.sub(r"\s+", "", item).strip()
        if clean and clean not in {"无", "不详", "未知", "unknown", "None"}:
            items.append(clean)
    return list(dict.fromkeys(items))


def hhi(counter: Counter) -> float:
    total = sum(counter.values())
    if not total:
        return 0.0
    return round(sum((count / total) ** 2 for count in counter.values()), 4)


def ratio(numerator: int | float, denominator: int | float) -> float:
    return round(float(numerator) / float(denominator), 4) if denominator else 0.0


def edge_key(left: str, right: str) -> str:
    first, second = sorted([left, right])
    return f"{first} -- {second}"


def evidence_strength(record_count: int, author_missing_rate: float, institution_missing_rate: float) -> str:
    if record_count >= 100 and author_missing_rate <= 0.2 and institution_missing_rate <= 0.2:
        return "strong"
    if record_count >= 50 and author_missing_rate <= 0.3 and institution_missing_rate <= 0.3:
        return "medium"
    return "weak"


def main() -> int:
    args = parse_args()
    records = load_records(Path(args.input))

    author_counts: Counter[str] = Counter()
    institution_counts: Counter[str] = Counter()
    author_edges: Counter[str] = Counter()
    institution_edges: Counter[str] = Counter()
    yearly_authors: dict[str, set[str]] = defaultdict(set)
    seen_authors: set[str] = set()
    yearly_new_author_ratio: dict[str, float] = {}
    multi_author_records = 0
    cross_institution_records = 0
    missing_author = 0
    missing_institution = 0

    yearly_author_sets: dict[str, list[set[str]]] = defaultdict(list)

    for row in records:
        authors = split_people(pick(row, "authors", "author", "作者", "第一作者"))
        institutions = split_people(pick(row, "institutions", "institution", "affiliation", "单位", "机构"))
        year = pick(row, "year", "年份") or "unknown"

        if not authors:
            missing_author += 1
        if not institutions:
            missing_institution += 1

        author_counts.update(authors)
        institution_counts.update(institutions)
        yearly_authors[year].update(authors)
        yearly_author_sets[year].append(set(authors))

        if len(authors) >= 2:
            multi_author_records += 1
            for left, right in itertools.combinations(authors, 2):
                author_edges[edge_key(left, right)] += 1

        unique_institutions = list(dict.fromkeys(institutions))
        if len(unique_institutions) >= 2:
            cross_institution_records += 1
            for left, right in itertools.combinations(unique_institutions, 2):
                institution_edges[edge_key(left, right)] += 1

    for year in sorted(yearly_authors):
        authors = yearly_authors[year]
        if year == "unknown":
            continue
        new_authors = authors - seen_authors
        yearly_new_author_ratio[year] = ratio(len(new_authors), len(authors))
        seen_authors.update(authors)

    author_missing_rate = ratio(missing_author, len(records))
    institution_missing_rate = ratio(missing_institution, len(records))
    strength = evidence_strength(len(records), author_missing_rate, institution_missing_rate)
    warnings = []
    if len(records) < 50:
        warnings.append("样本低于 50 条，作者机构网络只能作为初步观察。")
    if author_missing_rate > 0.3:
        warnings.append("作者字段缺失率高于 30%，不得给作者集中度高置信判断。")
    if institution_missing_rate > 0.3:
        warnings.append("机构字段缺失率高于 30%，不得给机构生态高置信判断。")

    stats = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "record_count": len(records),
        "author_count": len(author_counts),
        "author_mention_count": sum(author_counts.values()),
        "institution_count": len(institution_counts),
        "author_missing_rate": author_missing_rate,
        "institution_missing_rate": institution_missing_rate,
        "multi_author_record_ratio": ratio(multi_author_records, len(records)),
        "cross_institution_record_ratio": ratio(cross_institution_records, len(records)),
        "top_authors": author_counts.most_common(20),
        "top_institutions": institution_counts.most_common(20),
        "author_concentration_top10": ratio(sum(count for _, count in author_counts.most_common(10)), sum(author_counts.values())),
        "institution_hhi": hhi(institution_counts),
        "author_collaboration_edges": author_edges.most_common(30),
        "institution_collaboration_edges": institution_edges.most_common(30),
        "yearly_new_author_ratio": yearly_new_author_ratio,
        "evidence_strength": strength,
        "warnings": warnings,
    }

    lines = [
        "# 作者机构网络分析报告",
        "",
        f"- 样本数：{stats['record_count']}",
        f"- 作者数：{stats['author_count']}",
        f"- 作者署名次数：{stats['author_mention_count']}",
        f"- 机构数：{stats['institution_count']}",
        f"- 作者缺失率：{stats['author_missing_rate']}",
        f"- 机构缺失率：{stats['institution_missing_rate']}",
        f"- 多作者题录比例：{stats['multi_author_record_ratio']}",
        f"- 跨机构题录比例：{stats['cross_institution_record_ratio']}",
        f"- 前 10 作者署名集中度：{stats['author_concentration_top10']}",
        f"- 机构 HHI：{stats['institution_hhi']}",
        f"- 证据强度：{stats['evidence_strength']}",
        "",
        "## 降级提示",
    ]
    lines.extend([f"- {item}" for item in warnings] or ["- 暂无阻断性降级提示。"])
    lines.append("")
    lines.append("## 高频作者")
    lines.extend([f"- {name}：{count}" for name, count in stats["top_authors"][:10]])
    lines.append("")
    lines.append("## 高频机构")
    lines.extend([f"- {name}：{count}" for name, count in stats["top_institutions"][:10]])
    lines.append("")
    lines.append("## 作者合作边")
    lines.extend([f"- {edge}：{count}" for edge, count in stats["author_collaboration_edges"][:10]] or ["- 未形成可统计合作边。"])
    lines.append("")
    lines.append("## 机构合作边")
    lines.extend([f"- {edge}：{count}" for edge, count in stats["institution_collaboration_edges"][:10]] or ["- 未形成可统计机构合作边。"])
    lines.append("")
    lines.append("## 年度新作者比例")
    lines.extend([f"- {year}：{value}" for year, value in stats["yearly_new_author_ratio"].items()] or ["- 年份信息不足。"])

    Path(args.output_json).write_text(json.dumps(stats, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    Path(args.output_md).write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(json.dumps(stats, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
