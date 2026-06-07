#!/usr/bin/env python3
"""Generate journal quantitative analysis report from title/PDF/RAG inputs."""

from __future__ import annotations

import argparse
import csv
import json
import math
import re
from collections import Counter
from datetime import datetime
from pathlib import Path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate journal quantitative analysis report.")
    parser.add_argument("--title-list", required=True, help="CSV/JSON/XLSX title list.")
    parser.add_argument("--pdf-report", default="", help="PDF check report text/JSON.")
    parser.add_argument("--rag-report", default="", help="RAG import handoff text/JSON.")
    parser.add_argument("--expected-title-count", type=int, default=0, help="Expected total title count.")
    parser.add_argument("--expected-years", default="", help="Comma-separated expected years.")
    parser.add_argument("--output-md", required=True, help="Output Markdown report.")
    parser.add_argument("--output-json", required=True, help="Output JSON statistics.")
    return parser.parse_args()


def load_records(path: Path) -> list[dict]:
    suffix = path.suffix.lower()
    if suffix == ".json":
        data = json.loads(path.read_text(encoding="utf-8"))
        return list(data if isinstance(data, list) else data.get("records", []))
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as fh:
            return list(csv.DictReader(fh))
    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except Exception as exc:
            raise SystemExit("openpyxl is required for xlsx input") from exc
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        headers = [str(cell).strip() if cell is not None else "" for cell in next(rows)]
        return [{headers[i]: row[i] for i in range(min(len(headers), len(row)))} for row in rows]
    raise SystemExit(f"Unsupported input format: {path.suffix}")


def pick(row: dict, *keys: str) -> str:
    lowered = {str(k).strip().lower(): v for k, v in row.items()}
    for key in keys:
        value = lowered.get(key.lower())
        if value is not None and str(value).strip():
            return str(value).strip()
    return ""


def parse_count_report(path_text: str, keys: list[str]) -> int | None:
    if not path_text:
        return None
    path = Path(path_text)
    if not path.exists():
        return None
    text = path.read_text(encoding="utf-8", errors="ignore")
    if path.suffix.lower() == ".json":
        data = json.loads(text)
        for key in keys:
            if key in data:
                try:
                    return int(data[key])
                except Exception:
                    pass
    for key in keys:
        patterns = [
            rf"{re.escape(key)}\s*[:：]\s*(\d+)",
            rf"{re.escape(key)}[^\d]{{0,8}}(\d+)",
        ]
        for pattern in patterns:
            match = re.search(pattern, text)
            if match:
                return int(match.group(1))
    return None


def ratio(numerator: int | float, denominator: int | float) -> float:
    return round(float(numerator) / float(denominator), 4) if denominator else 0.0


def coefficient_variation(counts: Counter) -> float:
    values = list(counts.values())
    if not values:
        return 0.0
    mean = sum(values) / len(values)
    if mean == 0:
        return 0.0
    variance = sum((value - mean) ** 2 for value in values) / len(values)
    return round(math.sqrt(variance) / mean, 4)


def grade(title_coverage: float, pdf_coverage: float, rag_availability: float, missing_rate: float) -> str:
    if title_coverage >= 0.8 and pdf_coverage >= 0.5 and rag_availability >= 0.5:
        return "high"
    if title_coverage >= 0.6 and missing_rate <= 0.3:
        return "medium"
    return "low"


def evidence_strength(data_quality_grade: str, pdf_coverage: float, rag_availability: float) -> str:
    if data_quality_grade == "high" and pdf_coverage >= 0.5 and rag_availability >= 0.5:
        return "strong"
    if data_quality_grade in {"high", "medium"} and pdf_coverage >= 0.3:
        return "medium"
    return "weak"


def main() -> int:
    args = parse_args()
    records = load_records(Path(args.title_list))
    title_count = len(records)
    years = Counter(pick(row, "year", "年份") for row in records if pick(row, "year", "年份"))
    columns = Counter(pick(row, "column", "栏目") for row in records if pick(row, "column", "栏目"))

    titles = [pick(row, "title", "题名") for row in records]
    duplicate_titles = len([title for title, count in Counter(titles).items() if title and count > 1])
    key_fields = ["title", "题名", "author", "作者", "institution", "单位", "abstract", "摘要", "keywords", "关键词"]
    missing_cells = 0
    checked_cells = 0
    for row in records:
        for left, right in [("title", "题名"), ("author", "作者"), ("institution", "单位"), ("abstract", "摘要"), ("keywords", "关键词")]:
            checked_cells += 1
            if not pick(row, left, right):
                missing_cells += 1

    expected_title_count = args.expected_title_count or title_count
    expected_years = [item.strip() for item in args.expected_years.split(",") if item.strip()]
    title_coverage = ratio(title_count, expected_title_count)
    year_coverage = ratio(len(years), len(expected_years)) if expected_years else 1.0

    pdf_count = parse_count_report(args.pdf_report, ["pdf_count", "PDF 数", "PDF数量", "PDF 总数"])
    rag_count = parse_count_report(args.rag_report, ["rag_doc_count", "RAG 文档数", "入库条目数量", "成功数量"])
    pdf_count = pdf_count if pdf_count is not None else sum(1 for row in records if pick(row, "pdf_status", "PDF 状态") in {"有", "yes", "true", "1", "complete"})
    rag_count = rag_count if rag_count is not None else 0

    pdf_coverage = ratio(pdf_count, title_count)
    rag_availability = ratio(rag_count, pdf_count)
    missing_rate = ratio(missing_cells, checked_cells)
    data_quality_grade = grade(title_coverage, pdf_coverage, rag_availability, missing_rate)
    strength = evidence_strength(data_quality_grade, pdf_coverage, rag_availability)
    sample_coverage_rate = round(title_coverage * 0.4 + pdf_coverage * 0.3 + rag_availability * 0.3, 4)

    warnings = []
    if title_count < 50:
        warnings.append("题录数低于 50，题名风格分析只能作为初步观察。")
    if year_coverage < 0.8:
        warnings.append("年度覆盖率低于 80%，选题趋势判断存在盲区。")
    if pdf_coverage < 0.3:
        warnings.append("PDF 覆盖率低于 30%，后续论证风格分析为低置信度。")
    if rag_availability < 0.5:
        warnings.append("RAG 可用率低于 50%，论证风格和参考文献生态不得给高置信结论。")

    stats = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "title_count": title_count,
        "expected_title_count": expected_title_count,
        "title_coverage_rate": title_coverage,
        "year_coverage_rate": year_coverage,
        "pdf_count": pdf_count,
        "pdf_coverage_rate": pdf_coverage,
        "rag_doc_count": rag_count,
        "rag_availability_rate": rag_availability,
        "sample_coverage_rate": sample_coverage_rate,
        "year_distribution_cv": coefficient_variation(years),
        "column_distribution_cv": coefficient_variation(columns),
        "duplicate_title_rate": ratio(duplicate_titles, title_count),
        "missing_field_rate": missing_rate,
        "data_quality_grade": data_quality_grade,
        "evidence_strength": strength,
        "warnings": warnings,
        "year_counts": dict(years.most_common()),
        "column_counts": dict(columns.most_common()),
    }

    lines = [
        "# 数据性量化分析报告",
        "",
        f"- 题录数：{title_count}",
        f"- 题录覆盖率：{title_coverage}",
        f"- 年度覆盖率：{year_coverage}",
        f"- PDF 覆盖率：{pdf_coverage}",
        f"- RAG 可用率：{rag_availability}",
        f"- 综合样本覆盖率：{sample_coverage_rate}",
        f"- 数据质量等级：{data_quality_grade}",
        f"- 证据强度：{strength}",
        "",
        "## 数据质量警告",
    ]
    lines.extend([f"- {item}" for item in warnings] or ["- 暂无阻断性警告。"])
    lines.append("")
    lines.append("## 年度分布")
    lines.extend([f"- {key}：{value}" for key, value in stats["year_counts"].items()])
    lines.append("")
    lines.append("## 栏目分布")
    lines.extend([f"- {key}：{value}" for key, value in stats["column_counts"].items()])

    Path(args.output_json).write_text(json.dumps(stats, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    Path(args.output_md).write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(json.dumps(stats, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

