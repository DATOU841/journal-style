#!/usr/bin/env python3
"""Analyze public author profile, byline order, and funding metadata."""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter
from datetime import datetime
from pathlib import Path


SPLIT_PATTERN = re.compile(r"[;；,，、/|]+")
UNKNOWN_VALUES = {"", "无", "不详", "未知", "unknown", "none", "None", "null"}
SENIOR_TITLE_PATTERN = re.compile(r"教授|副教授|研究员|副研究员|博导|硕导|导师")
PROFESSOR_PATTERN = re.compile(r"教授|研究员|博导|硕导|导师")
ASSOCIATE_PROFESSOR_PATTERN = re.compile(r"副教授|副研究员")
LECTURER_PATTERN = re.compile(r"讲师|助理研究员")
MASTER_PATTERN = re.compile(r"硕士|硕士生|硕士研究生|硕士在读")
PHD_PATTERN = re.compile(r"博士|博士生|博士研究生|博士在读")
UNDERGRAD_PATTERN = re.compile(r"本科|本科生")
STUDENT_PATTERN = re.compile(r"硕士|博士|研究生|本科生|学生")
ADVISOR_RELATION_PATTERN = re.compile(r"导师|指导教师|师生|指导关系|指导老师")
NATIONAL_FUND_PATTERN = re.compile(r"国家社科|国家自然|国家艺术|国家级|国家社会科学|国家自然科学|教育部重大|重大项目")
PROVINCIAL_FUND_PATTERN = re.compile(r"省社科|省级|省部级|教育部|文化和旅游部|中国文联|全国教育科学")
MUNICIPAL_FUND_PATTERN = re.compile(r"市厅|市级|厅级|地市")
UNIVERSITY_FUND_PATTERN = re.compile(r"校级|高校|学院|大学|研究生创新")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Analyze public author profile and byline metadata.")
    parser.add_argument("--input", required=True, help="CSV/JSON/XLSX title metadata.")
    parser.add_argument("--output-json", default="journal-author-profile-and-byline-statistics.json")
    parser.add_argument("--output-md", default="journal-author-profile-and-byline-report.md")
    return parser.parse_args()


def load_records(path: Path) -> list[dict]:
    suffix = path.suffix.lower()
    if suffix == ".json":
        data = json.loads(path.read_text(encoding="utf-8"))
        return list(data if isinstance(data, list) else data.get("records", []))
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as fh:
            return list(csv.DictReader(fh))
    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except Exception as exc:  # pragma: no cover - dependency guard
            raise SystemExit("openpyxl is required for xlsx input") from exc
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        headers = [str(cell).strip() if cell is not None else "" for cell in next(rows)]
        return [{headers[i]: row[i] for i in range(min(len(headers), len(row)))} for row in rows]
    raise SystemExit(f"Unsupported input format: {path.suffix}")


def pick(row: dict, *keys: str) -> str:
    lowered = {str(k).strip().lower(): v for k, v in row.items()}
    for key in keys:
        value = lowered.get(key.lower())
        if value is not None and str(value).strip():
            return str(value).strip()
    return ""


def split_items(value: str) -> list[str]:
    items = []
    for item in SPLIT_PATTERN.split(value or ""):
        clean = re.sub(r"\s+", "", str(item)).strip()
        if clean and clean not in UNKNOWN_VALUES:
            items.append(clean)
    return items


def ratio(numerator: int | float, denominator: int | float) -> float:
    return round(float(numerator) / float(denominator), 4) if denominator else 0.0


def classify_title(value: str) -> str:
    text = value or ""
    if ASSOCIATE_PROFESSOR_PATTERN.search(text):
        return "副教授/副研究员"
    if PROFESSOR_PATTERN.search(text):
        return "教授/研究员/导师"
    if LECTURER_PATTERN.search(text):
        return "讲师/助理研究员"
    if STUDENT_PATTERN.search(text):
        return "学生"
    return "未知" if not text else "其他"


def classify_stage(value: str) -> str:
    text = value or ""
    if PHD_PATTERN.search(text):
        return "博士生/博士"
    if MASTER_PATTERN.search(text):
        return "硕士生/硕士"
    if UNDERGRAD_PATTERN.search(text):
        return "本科生"
    if SENIOR_TITLE_PATTERN.search(text):
        return "高年资教师/研究员"
    return "未知" if not text else "其他"


def classify_fund(value: str) -> str:
    text = value or ""
    if not text or text in UNKNOWN_VALUES:
        return "无/未标注"
    if NATIONAL_FUND_PATTERN.search(text):
        return "国家级/重大"
    if PROVINCIAL_FUND_PATTERN.search(text):
        return "省部级"
    if MUNICIPAL_FUND_PATTERN.search(text):
        return "市厅级"
    if UNIVERSITY_FUND_PATTERN.search(text):
        return "校级/高校"
    return "其他"


def aligned_value(values: list[str], index: int) -> str:
    return values[index] if index < len(values) else ""


def contains_student(*values: str) -> bool:
    return any(STUDENT_PATTERN.search(value or "") for value in values)


def is_senior(*values: str) -> bool:
    return any(SENIOR_TITLE_PATTERN.search(value or "") for value in values)


def is_master(*values: str) -> bool:
    return any(MASTER_PATTERN.search(value or "") for value in values)


def is_phd(*values: str) -> bool:
    return any(PHD_PATTERN.search(value or "") for value in values)


def evidence_strength(record_count: int, title_coverage: float, stage_coverage: float, fund_coverage: float) -> str:
    if record_count >= 100 and title_coverage >= 0.7 and stage_coverage >= 0.7 and fund_coverage >= 0.7:
        return "strong"
    if record_count >= 50 and title_coverage >= 0.5 and stage_coverage >= 0.5:
        return "medium"
    return "weak"


def main() -> int:
    args = parse_args()
    records = load_records(Path(args.input))

    author_count_distribution: Counter[str] = Counter()
    first_author_titles: Counter[str] = Counter()
    second_author_stages: Counter[str] = Counter()
    corresponding_positions: Counter[str] = Counter()
    fund_levels: Counter[str] = Counter()
    missing_authors = 0
    records_with_titles = 0
    records_with_stage_or_degree = 0
    records_with_fund_field = 0
    funded_records = 0
    student_author_records = 0
    senior_first_student_second = 0
    senior_first_master_second = 0
    senior_first_phd_second = 0
    explicit_advisor_student = 0
    corresponding_records = 0
    aligned_profile_records = 0
    sample_flags: list[dict] = []

    for row in records:
        authors = split_items(pick(row, "authors", "author", "作者", "第一作者"))
        titles = split_items(pick(row, "author_titles", "academic_title", "职称", "作者职称"))
        degrees = split_items(pick(row, "author_degrees", "degree", "学位", "作者学位"))
        stages = split_items(pick(row, "author_stages", "author_stage", "author_roles", "作者身份", "身份"))
        corresponding = split_items(pick(row, "corresponding_author", "通讯作者"))
        fund = pick(row, "fund", "funding", "基金", "基金项目", "fund_project")
        relation = pick(row, "advisor_student_relation", "师生关系说明", "导师学生说明")
        title = pick(row, "title", "题名")

        if not authors:
            missing_authors += 1
        author_count = len(authors)
        if author_count <= 1:
            author_count_distribution["单作者"] += 1
        elif author_count == 2:
            author_count_distribution["双作者"] += 1
        else:
            author_count_distribution["三人及以上"] += 1

        if titles:
            records_with_titles += 1
        if stages or degrees:
            records_with_stage_or_degree += 1
        if fund or "fund" in {str(key).lower() for key in row} or "基金" in {str(key) for key in row}:
            records_with_fund_field += 1

        first_title = aligned_value(titles, 0)
        first_stage = aligned_value(stages, 0)
        first_degree = aligned_value(degrees, 0)
        second_title = aligned_value(titles, 1)
        second_stage = aligned_value(stages, 1)
        second_degree = aligned_value(degrees, 1)

        first_author_titles[classify_title(" ".join([first_title, first_stage, first_degree]))] += 1
        if author_count >= 2:
            second_author_stages[classify_stage(" ".join([second_title, second_stage, second_degree]))] += 1

        if author_count and (len(titles) == author_count or len(stages) == author_count or len(degrees) == author_count):
            aligned_profile_records += 1

        all_profile_values = titles + stages + degrees
        if contains_student(*all_profile_values):
            student_author_records += 1

        senior_first = is_senior(first_title, first_stage, first_degree)
        master_second = is_master(second_title, second_stage, second_degree)
        phd_second = is_phd(second_title, second_stage, second_degree)
        student_second = contains_student(second_title, second_stage, second_degree)
        if author_count >= 2 and senior_first and student_second:
            senior_first_student_second += 1
            if len(sample_flags) < 8:
                sample_flags.append({
                    "title": title,
                    "flag": "高年资一作 + 学生二作线索",
                    "first_author": authors[0] if authors else "",
                    "second_author": authors[1] if len(authors) >= 2 else "",
                })
        if author_count >= 2 and senior_first and master_second:
            senior_first_master_second += 1
        if author_count >= 2 and senior_first and phd_second:
            senior_first_phd_second += 1

        if ADVISOR_RELATION_PATTERN.search(relation or ""):
            explicit_advisor_student += 1

        if corresponding:
            corresponding_records += 1
            for person in corresponding:
                if person in authors:
                    corresponding_positions[f"第{authors.index(person) + 1}作者"] += 1
                else:
                    corresponding_positions["未能匹配作者序"] += 1

        fund_level = classify_fund(fund)
        fund_levels[fund_level] += 1
        if fund_level != "无/未标注":
            funded_records += 1

    record_count = len(records)
    title_coverage = ratio(records_with_titles, record_count)
    stage_coverage = ratio(records_with_stage_or_degree, record_count)
    fund_field_coverage = ratio(records_with_fund_field, record_count)
    author_missing_rate = ratio(missing_authors, record_count)
    strength = evidence_strength(record_count, title_coverage, stage_coverage, fund_field_coverage)
    warnings = []
    if record_count < 50:
        warnings.append("样本低于 50 条，作者公共身份与署名基金结构只能作为初步观察。")
    if title_coverage < 0.5:
        warnings.append("职称字段覆盖率低于 50%，不得给第一作者职称结构强结论。")
    if stage_coverage < 0.5:
        warnings.append("学位/身份字段覆盖率低于 50%，不得给学生作者参与度强结论。")
    if fund_field_coverage < 0.5:
        warnings.append("基金字段覆盖率低于 50%，基金结构只能写题录层初步统计。")
    if explicit_advisor_student == 0 and senior_first_master_second:
        warnings.append("存在高年资一作 + 硕士二作线索，但无明确师生关系说明，不得写成确定的“导一硕二”。")

    stats = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "record_count": record_count,
        "author_missing_rate": author_missing_rate,
        "author_title_coverage_rate": title_coverage,
        "author_stage_or_degree_coverage_rate": stage_coverage,
        "fund_field_coverage_rate": fund_field_coverage,
        "aligned_author_profile_rate": ratio(aligned_profile_records, record_count),
        "author_count_distribution": dict(author_count_distribution),
        "first_author_title_distribution": dict(first_author_titles),
        "second_author_stage_distribution": dict(second_author_stages),
        "student_author_record_ratio": ratio(student_author_records, record_count),
        "senior_first_student_second_ratio": ratio(senior_first_student_second, record_count),
        "senior_first_master_second_ratio": ratio(senior_first_master_second, record_count),
        "senior_first_phd_second_ratio": ratio(senior_first_phd_second, record_count),
        "explicit_advisor_student_relation_ratio": ratio(explicit_advisor_student, record_count),
        "corresponding_author_record_ratio": ratio(corresponding_records, record_count),
        "corresponding_author_position_distribution": dict(corresponding_positions),
        "funded_record_ratio": ratio(funded_records, record_count),
        "fund_level_distribution": dict(fund_levels),
        "sample_public_identity_flags": sample_flags,
        "evidence_strength": strength,
        "warnings": warnings,
    }

    lines = [
        "# 作者公共身份与署名基金结构分析报告",
        "",
        f"- 样本数：{stats['record_count']}",
        f"- 作者字段缺失率：{stats['author_missing_rate']}",
        f"- 职称字段覆盖率：{stats['author_title_coverage_rate']}",
        f"- 学位/身份字段覆盖率：{stats['author_stage_or_degree_coverage_rate']}",
        f"- 基金字段覆盖率：{stats['fund_field_coverage_rate']}",
        f"- 作者身份字段可对齐率：{stats['aligned_author_profile_rate']}",
        f"- 学生作者参与比例：{stats['student_author_record_ratio']}",
        f"- 高年资一作 + 学生二作线索比例：{stats['senior_first_student_second_ratio']}",
        f"- 高年资一作 + 硕士二作线索比例：{stats['senior_first_master_second_ratio']}",
        f"- 高年资一作 + 博士二作线索比例：{stats['senior_first_phd_second_ratio']}",
        f"- 明确师生关系说明比例：{stats['explicit_advisor_student_relation_ratio']}",
        f"- 基金论文比例：{stats['funded_record_ratio']}",
        f"- 证据强度：{stats['evidence_strength']}",
        "",
        "## 降级提示",
    ]
    lines.extend([f"- {item}" for item in warnings] or ["- 暂无阻断性降级提示。"])
    lines.extend(["", "## 作者数结构"])
    lines.extend([f"- {name}：{count}" for name, count in author_count_distribution.items()] or ["- 无可统计作者数结构。"])
    lines.extend(["", "## 第一作者职称分布"])
    lines.extend([f"- {name}：{count}" for name, count in first_author_titles.items()] or ["- 职称字段不足。"])
    lines.extend(["", "## 第二作者身份分布"])
    lines.extend([f"- {name}：{count}" for name, count in second_author_stages.items()] or ["- 第二作者身份字段不足。"])
    lines.extend(["", "## 通讯作者位置"])
    lines.extend([f"- {name}：{count}" for name, count in corresponding_positions.items()] or ["- 通讯作者字段不足或未标注。"])
    lines.extend(["", "## 基金层级分布"])
    lines.extend([f"- {name}：{count}" for name, count in fund_levels.items()] or ["- 基金字段不足。"])
    lines.extend(["", "## 公开身份组合线索样本"])
    lines.extend([
        f"- {item['title']}：{item['flag']}（{item['first_author']} / {item['second_author']}）"
        for item in sample_flags
    ] or ["- 暂无可列样本。"])
    lines.extend([
        "",
        "## 不得推断项",
        "- 不得把高年资一作 + 硕士/博士二作线索直接写成确定的导师一作、学生二作，除非公开材料明确说明。",
        "- 不得用作者排位、职称、学位或基金信息推断录用倾向、私人关系或投稿捷径。",
    ])

    Path(args.output_json).write_text(json.dumps(stats, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    Path(args.output_md).write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(json.dumps(stats, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
