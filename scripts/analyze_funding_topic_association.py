#!/usr/bin/env python3
"""Analyze associations between funding levels and journal topic patterns."""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path


NATIONAL_FUND_PATTERN = re.compile(r"国家社科|国家自然|国家艺术|国家级|国家社会科学|国家自然科学|教育部重大|重大项目")
PROVINCIAL_FUND_PATTERN = re.compile(r"省社科|省级|省部级|教育部|文化和旅游部|中国文联|全国教育科学")
MUNICIPAL_FUND_PATTERN = re.compile(r"市厅|市级|厅级|地市")
UNIVERSITY_FUND_PATTERN = re.compile(r"校级|高校|学院|大学|研究生创新")
SPLIT_PATTERN = re.compile(r"[;；,，、\s/|]+")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Analyze funding-topic association from title metadata.")
    parser.add_argument("--input", required=True, help="CSV/JSON/XLSX title metadata.")
    parser.add_argument("--user-keywords", default="", help="Optional comma-separated user keywords.")
    parser.add_argument("--output-json", default="journal-funding-topic-association-statistics.json")
    parser.add_argument("--output-md", default="journal-funding-topic-association-report.md")
    return parser.parse_args()


def load_records(path: Path) -> list[dict]:
    suffix = path.suffix.lower()
    if suffix == ".json":
        data = json.loads(path.read_text(encoding="utf-8"))
        return list(data if isinstance(data, list) else data.get("records", []))
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as fh:
            return list(csv.DictReader(fh))
    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except Exception as exc:  # pragma: no cover - dependency guard
            raise SystemExit("openpyxl is required for xlsx input") from exc
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        headers = [str(cell).strip() if cell is not None else "" for cell in next(rows)]
        return [{headers[i]: row[i] for i in range(min(len(headers), len(row)))} for row in rows]
    raise SystemExit(f"Unsupported input format: {path.suffix}")


def pick(row: dict, *keys: str) -> str:
    lowered = {str(k).strip().lower(): v for k, v in row.items()}
    for key in keys:
        value = lowered.get(key.lower())
        if value is not None and str(value).strip():
            return str(value).strip()
    return ""


def classify_fund(value: str, explicit_level: str = "") -> str:
    text = explicit_level or value or ""
    if not text or text.strip() in {"无", "未标注", "不详", "未知", "none", "None"}:
        return "无/未标注"
    if NATIONAL_FUND_PATTERN.search(text):
        return "国家级/重大"
    if PROVINCIAL_FUND_PATTERN.search(text):
        return "省部级"
    if MUNICIPAL_FUND_PATTERN.search(text):
        return "市厅级"
    if UNIVERSITY_FUND_PATTERN.search(text):
        return "校级/高校"
    return "其他"


def split_items(value: str) -> list[str]:
    items = []
    for item in SPLIT_PATTERN.split(value or ""):
        clean = item.strip()
        if clean and clean not in {"无", "未知", "不详"}:
            items.append(clean)
    return items


def extract_keywords(row: dict) -> list[str]:
    values = [
        pick(row, "keywords", "keyword", "关键词", "关键字"),
        pick(row, "topic", "主题"),
    ]
    keywords = []
    for value in values:
        keywords.extend(split_items(value))
    if not keywords:
        title = pick(row, "title", "题名")
        abstract = pick(row, "abstract", "摘要")
        for token in split_items(" ".join([title, abstract])):
            if 2 <= len(token) <= 12:
                keywords.append(token)
    return list(dict.fromkeys(keywords))


def ratio(numerator: int | float, denominator: int | float) -> float:
    return round(float(numerator) / float(denominator), 4) if denominator else 0.0


def top_dict(counter: Counter, limit: int = 20) -> list[dict]:
    return [{"name": name, "count": count} for name, count in counter.most_common(limit)]


def main() -> int:
    args = parse_args()
    records = load_records(Path(args.input))
    user_keywords = split_items(args.user_keywords.replace(",", " "))

    fund_level_counts: Counter[str] = Counter()
    fund_keyword_counts: dict[str, Counter[str]] = defaultdict(Counter)
    fund_column_counts: dict[str, Counter[str]] = defaultdict(Counter)
    fund_material_counts: dict[str, Counter[str]] = defaultdict(Counter)
    fund_method_counts: dict[str, Counter[str]] = defaultdict(Counter)
    funded_keyword_counts: Counter[str] = Counter()
    nonfunded_keyword_counts: Counter[str] = Counter()
    user_keyword_hits: dict[str, Counter[str]] = {keyword: Counter() for keyword in user_keywords}
    records_with_fund_field = 0
    funded_records = 0

    for row in records:
        fund = pick(row, "fund", "funding", "基金", "基金项目", "fund_project")
        explicit_level = pick(row, "fund_level", "基金层级")
        if fund or explicit_level or "fund" in {str(key).lower() for key in row} or "基金" in {str(key) for key in row}:
            records_with_fund_field += 1
        level = classify_fund(fund, explicit_level)
        fund_level_counts[level] += 1
        is_funded = level != "无/未标注"
        if is_funded:
            funded_records += 1

        keywords = extract_keywords(row)
        column = pick(row, "column", "栏目") or "未标栏目"
        materials = split_items(pick(row, "material_type", "materials", "材料类型"))
        methods = split_items(pick(row, "method_type", "methods", "方法类型"))

        fund_keyword_counts[level].update(keywords)
        fund_column_counts[level][column] += 1
        fund_material_counts[level].update(materials or ["未标材料"])
        fund_method_counts[level].update(methods or ["未标方法"])
        if is_funded:
            funded_keyword_counts.update(keywords)
        else:
            nonfunded_keyword_counts.update(keywords)
        searchable = " ".join([pick(row, "title", "题名"), pick(row, "abstract", "摘要"), pick(row, "keywords", "关键词")])
        for keyword in user_keywords:
            if keyword and keyword in searchable:
                user_keyword_hits[keyword][level] += 1

    record_count = len(records)
    fund_coverage = ratio(records_with_fund_field, record_count)
    funded_ratio = ratio(funded_records, record_count)
    level_details = {}
    for level in fund_level_counts:
        level_details[level] = {
            "record_count": fund_level_counts[level],
            "top_keywords": top_dict(fund_keyword_counts[level], 15),
            "top_columns": top_dict(fund_column_counts[level], 10),
            "top_materials": top_dict(fund_material_counts[level], 10),
            "top_methods": top_dict(fund_method_counts[level], 10),
        }

    funded_only_keywords = []
    for keyword, count in funded_keyword_counts.most_common(30):
        funded_only_keywords.append({
            "keyword": keyword,
            "funded_count": count,
            "nonfunded_count": nonfunded_keyword_counts.get(keyword, 0),
        })

    warnings = []
    if record_count < 50:
        warnings.append("样本低于 50 条，基金与选题关联只能作为初步观察。")
    if fund_coverage < 0.5:
        warnings.append("基金字段覆盖率低于 50%，不得给基金偏好强结论。")
    for level, count in fund_level_counts.items():
        if level != "无/未标注" and count < 5:
            warnings.append(f"{level} 样本低于 5 条，只能列线索。")
    strength = "strong" if record_count >= 100 and fund_coverage >= 0.7 else "medium" if record_count >= 50 and fund_coverage >= 0.5 else "weak"

    stats = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "record_count": record_count,
        "fund_field_coverage_rate": fund_coverage,
        "funded_record_ratio": funded_ratio,
        "fund_level_counts": dict(fund_level_counts),
        "fund_level_details": level_details,
        "funded_vs_nonfunded_keywords": funded_only_keywords,
        "user_keyword_hits_by_fund_level": {key: dict(value) for key, value in user_keyword_hits.items()},
        "evidence_strength": strength,
        "warnings": warnings,
    }

    lines = [
        "# 基金与选题关联分析报告",
        "",
        f"- 样本数：{record_count}",
        f"- 基金字段覆盖率：{fund_coverage}",
        f"- 基金论文比例：{funded_ratio}",
        f"- 证据强度：{strength}",
        "",
        "## 降级提示",
    ]
    lines.extend([f"- {item}" for item in warnings] or ["- 暂无阻断性降级提示。"])
    lines.extend(["", "## 基金层级分布"])
    lines.extend([f"- {level}：{count}" for level, count in fund_level_counts.items()])
    for level, item in level_details.items():
        lines.extend(["", f"## {level}"])
        lines.append("- 高频关键词：" + "；".join([f"{entry['name']}({entry['count']})" for entry in item["top_keywords"][:10]]) if item["top_keywords"] else "- 高频关键词：无")
        lines.append("- 栏目分布：" + "；".join([f"{entry['name']}({entry['count']})" for entry in item["top_columns"][:8]]) if item["top_columns"] else "- 栏目分布：无")
        lines.append("- 材料分布：" + "；".join([f"{entry['name']}({entry['count']})" for entry in item["top_materials"][:8]]) if item["top_materials"] else "- 材料分布：无")
        lines.append("- 方法分布：" + "；".join([f"{entry['name']}({entry['count']})" for entry in item["top_methods"][:8]]) if item["top_methods"] else "- 方法分布：无")
    lines.extend(["", "## 用户关键词命中"])
    lines.extend([f"- {keyword}：{dict(counter)}" for keyword, counter in user_keyword_hits.items()] or ["- 未提供用户关键词。"])
    lines.extend([
        "",
        "## 不得推断项",
        "- 不得写“有基金更容易录用”或“无基金不适合投稿”。",
        "- 只能写已刊样本中某类基金论文更常见于某些栏目、关键词、材料或方法。",
    ])

    Path(args.output_json).write_text(json.dumps(stats, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    Path(args.output_md).write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(json.dumps(stats, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
