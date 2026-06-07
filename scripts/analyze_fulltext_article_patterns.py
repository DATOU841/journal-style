#!/usr/bin/env python3
"""Analyze post-RAG/full-text patterns from extracted article records."""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter
from datetime import datetime
from pathlib import Path


SECTION_PATTERN = re.compile(r"^(#{1,3}\s*)?((第?[一二三四五六七八九十]+[章节、.．])|(\d+(\.\d+)*[、.．]))\s*(.+)$", re.MULTILINE)
ABSTRACT_PATTERN = re.compile(r"摘要[:：]?\s*(.*?)(关键词|关键字)[:：]", re.S)
KEYWORD_PATTERN = re.compile(r"(关键词|关键字)[:：]\s*([^\n。]+)")
RECEIVED_PATTERN = re.compile(r"(收稿日期|收稿时间)[:：]?\s*([0-9]{4}[-年/.][0-9]{1,2}[-月/.][0-9]{1,2})")
REVISED_PATTERN = re.compile(r"(修回日期|修订日期|返修日期)[:：]?\s*([0-9]{4}[-年/.][0-9]{1,2}[-月/.][0-9]{1,2})")
ACCEPTED_PATTERN = re.compile(r"(录用日期|接受日期)[:：]?\s*([0-9]{4}[-年/.][0-9]{1,2}[-月/.][0-9]{1,2})")
PUBLISHED_PATTERN = re.compile(r"(刊发日期|出版日期|发表日期)[:：]?\s*([0-9]{4}[-年/.][0-9]{1,2}[-月/.][0-9]{1,2})")
REFERENCE_PATTERN = re.compile(r"(\[\d+\]|［\d+］|^\d+[.．])", re.MULTILINE)
TABLE_PATTERN = re.compile(r"(表\s*\d+|表[一二三四五六七八九十]+)")
FIGURE_PATTERN = re.compile(r"(图\s*\d+|图[一二三四五六七八九十]+)")
NOTE_PATTERN = re.compile(r"(注释|脚注|尾注|注：)")

MATERIAL_MARKERS = {
    "档案": re.compile(r"档案|馆藏|案卷|史料"),
    "碑帖": re.compile(r"碑帖|碑刻|拓本|墨迹|法帖"),
    "访谈/田野": re.compile(r"访谈|田野|调研|问卷"),
    "图像": re.compile(r"图像|图版|图式|视觉"),
    "数据库/统计": re.compile(r"数据库|数据集|统计|量化"),
    "文献汇编": re.compile(r"文献|汇编|目录|版本"),
}

METHOD_MARKERS = {
    "考辨": re.compile(r"考辨|考证|辨析|校勘"),
    "比较": re.compile(r"比较|对比|异同"),
    "个案": re.compile(r"个案|案例"),
    "阐释": re.compile(r"阐释|诠释|解读"),
    "计量": re.compile(r"计量|统计|词频|网络分析"),
    "图像分析": re.compile(r"图像分析|视觉分析|图式分析"),
    "田野": re.compile(r"田野|访谈|调研"),
    "文本分析": re.compile(r"文本分析|话语分析|语义"),
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Analyze post-RAG/full-text article patterns.")
    parser.add_argument("--input", required=True, help="Directory of .txt/.md files or CSV/JSON/XLSX records.")
    parser.add_argument("--output-json", default="journal-rag-fulltext-pattern-statistics.json")
    parser.add_argument("--output-md", default="journal-rag-fulltext-pattern-report.md")
    return parser.parse_args()


def load_records(path: Path) -> list[dict]:
    if path.is_dir():
        records = []
        for file_path in sorted(path.glob("*")):
            if file_path.suffix.lower() not in {".txt", ".md"}:
                continue
            records.append({
                "article_id": file_path.stem,
                "title": file_path.stem,
                "fulltext": file_path.read_text(encoding="utf-8", errors="ignore"),
                "source_path": str(file_path),
            })
        return records
    suffix = path.suffix.lower()
    if suffix == ".json":
        data = json.loads(path.read_text(encoding="utf-8"))
        return list(data if isinstance(data, list) else data.get("records", []))
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as fh:
            return list(csv.DictReader(fh))
    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except Exception as exc:  # pragma: no cover - dependency guard
            raise SystemExit("openpyxl is required for xlsx input") from exc
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        headers = [str(cell).strip() if cell is not None else "" for cell in next(rows)]
        return [{headers[i]: row[i] for i in range(min(len(headers), len(row)))} for row in rows]
    raise SystemExit(f"Unsupported input format: {path.suffix}")


def pick(row: dict, *keys: str) -> str:
    lowered = {str(k).strip().lower(): v for k, v in row.items()}
    for key in keys:
        value = lowered.get(key.lower())
        if value is not None and str(value).strip():
            return str(value).strip()
    return ""


def split_keywords(value: str) -> list[str]:
    return [item.strip() for item in re.split(r"[;；,，、\s]+", value or "") if item.strip()]


def parse_date(text: str) -> datetime | None:
    if not text:
        return None
    normalized = text.replace("年", "-").replace("月", "-").replace("日", "").replace("/", "-").replace(".", "-")
    parts = [part for part in normalized.split("-") if part]
    if len(parts) < 3:
        return None
    try:
        return datetime(int(parts[0]), int(parts[1]), int(parts[2]))
    except ValueError:
        return None


def date_value(pattern: re.Pattern, text: str) -> str:
    match = pattern.search(text)
    return match.group(2) if match else ""


def ratio(numerator: int | float, denominator: int | float) -> float:
    return round(float(numerator) / float(denominator), 4) if denominator else 0.0


def main() -> int:
    args = parse_args()
    records = load_records(Path(args.input))

    section_counts: Counter[str] = Counter()
    material_counts: Counter[str] = Counter()
    method_counts: Counter[str] = Counter()
    keyword_counts: Counter[str] = Counter()
    year_counts: Counter[str] = Counter()
    column_counts: Counter[str] = Counter()
    abstract_lengths: list[int] = []
    keyword_lengths: list[int] = []
    article_summaries: list[dict] = []
    date_records: list[dict] = []
    total_tables = 0
    total_figures = 0
    total_notes = 0
    total_references = 0

    for index, row in enumerate(records, start=1):
        text = pick(row, "fulltext", "text", "全文", "正文")
        abstract = pick(row, "abstract", "摘要")
        keywords_text = pick(row, "keywords", "关键词", "关键字")
        if not abstract:
            match = ABSTRACT_PATTERN.search(text)
            if match:
                abstract = re.sub(r"\s+", "", match.group(1))
        if not keywords_text:
            match = KEYWORD_PATTERN.search(text)
            if match:
                keywords_text = match.group(2)
        keywords = split_keywords(keywords_text)
        keyword_counts.update(keywords)
        keyword_lengths.append(len(keywords))
        if abstract:
            abstract_lengths.append(len(re.sub(r"\s+", "", abstract)))

        year = pick(row, "year", "年份") or "unknown"
        column = pick(row, "column", "栏目") or "unknown"
        year_counts[year] += 1
        column_counts[column] += 1

        headings = []
        for match in SECTION_PATTERN.finditer(text):
            heading = re.sub(r"\s+", "", match.group(0)).strip()
            if len(heading) <= 40:
                headings.append(heading)
                section_counts[heading] += 1

        materials = [name for name, pattern in MATERIAL_MARKERS.items() if pattern.search(text)]
        methods = [name for name, pattern in METHOD_MARKERS.items() if pattern.search(text)]
        material_counts.update(materials)
        method_counts.update(methods)
        table_count = len(TABLE_PATTERN.findall(text))
        figure_count = len(FIGURE_PATTERN.findall(text))
        note_count = len(NOTE_PATTERN.findall(text))
        reference_count = len(REFERENCE_PATTERN.findall(text))
        total_tables += table_count
        total_figures += figure_count
        total_notes += note_count
        total_references += reference_count

        received = date_value(RECEIVED_PATTERN, text)
        revised = date_value(REVISED_PATTERN, text)
        accepted = date_value(ACCEPTED_PATTERN, text)
        published = date_value(PUBLISHED_PATTERN, text)
        received_dt = parse_date(received)
        accepted_dt = parse_date(accepted)
        published_dt = parse_date(published)
        review_days = (accepted_dt - received_dt).days if received_dt and accepted_dt else None
        publication_lag_days = (published_dt - accepted_dt).days if accepted_dt and published_dt else None
        if received or revised or accepted or published:
            date_records.append({
                "article_id": pick(row, "article_id", "id") or f"article-{index}",
                "title": pick(row, "title", "题名") or f"article-{index}",
                "received": received,
                "revised": revised,
                "accepted": accepted,
                "published": published,
                "review_days": review_days,
                "publication_lag_days": publication_lag_days,
            })

        if len(article_summaries) < 12:
            article_summaries.append({
                "article_id": pick(row, "article_id", "id") or f"article-{index}",
                "title": pick(row, "title", "题名") or f"article-{index}",
                "heading_count": len(headings),
                "sample_headings": headings[:6],
                "materials": materials,
                "methods": methods,
                "table_count": table_count,
                "figure_count": figure_count,
                "reference_marker_count": reference_count,
            })

    record_count = len(records)
    review_values = [item["review_days"] for item in date_records if item.get("review_days") is not None]
    publication_lag_values = [item["publication_lag_days"] for item in date_records if item.get("publication_lag_days") is not None]
    warnings = []
    if record_count < 10:
        warnings.append("全文样本低于 10 篇，只能写全文样本观察。")
    elif record_count < 20:
        warnings.append("全文样本低于 20 篇，不得写成全刊稳定风格。")
    if not date_records:
        warnings.append("未识别收稿、修回、录用或刊发日期，不能计算审稿/见刊周期。")
    if not abstract_lengths:
        warnings.append("摘要字段或正文摘要段识别不足，摘要结构需人工复核。")

    strength = "strong" if record_count >= 30 else "medium" if record_count >= 20 else "weak"
    stats = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "record_count": record_count,
        "year_counts": dict(year_counts),
        "column_counts": dict(column_counts),
        "abstract_length_mean": round(sum(abstract_lengths) / len(abstract_lengths), 2) if abstract_lengths else None,
        "keyword_count_mean": round(sum(keyword_lengths) / len(keyword_lengths), 2) if keyword_lengths else None,
        "top_keywords": keyword_counts.most_common(30),
        "top_headings": section_counts.most_common(30),
        "material_marker_counts": dict(material_counts.most_common()),
        "method_marker_counts": dict(method_counts.most_common()),
        "table_count_total": total_tables,
        "figure_count_total": total_figures,
        "note_marker_total": total_notes,
        "reference_marker_total": total_references,
        "date_record_count": len(date_records),
        "review_days_values": review_values,
        "publication_lag_days_values": publication_lag_values,
        "date_records": date_records[:30],
        "article_summaries": article_summaries,
        "evidence_strength": strength,
        "warnings": warnings,
    }

    lines = [
        "# RAG/全文样本模式挖掘报告",
        "",
        f"- 全文样本数：{record_count}",
        f"- 年份分布：{stats['year_counts']}",
        f"- 栏目分布：{stats['column_counts']}",
        f"- 摘要平均长度：{stats['abstract_length_mean']}",
        f"- 关键词平均数量：{stats['keyword_count_mean']}",
        f"- 证据强度：{strength}",
        "",
        "## 降级提示",
    ]
    lines.extend([f"- {item}" for item in warnings] or ["- 暂无阻断性降级提示。"])
    lines.extend(["", "## 高频章节标题"])
    lines.extend([f"- {name}：{count}" for name, count in section_counts.most_common(12)] or ["- 未识别稳定章节标题。"])
    lines.extend(["", "## 材料标记"])
    lines.extend([f"- {name}：{count}" for name, count in material_counts.most_common()] or ["- 未识别材料标记。"])
    lines.extend(["", "## 方法标记"])
    lines.extend([f"- {name}：{count}" for name, count in method_counts.most_common()] or ["- 未识别方法标记。"])
    lines.extend(["", "## 图表、注释、参考文献线索"])
    lines.extend([
        f"- 表格标记总数：{total_tables}",
        f"- 图像标记总数：{total_figures}",
        f"- 注释标记总数：{total_notes}",
        f"- 参考文献序号标记总数：{total_references}",
    ])
    lines.extend(["", "## 收稿、录用与刊发日期线索"])
    if date_records:
        for item in date_records[:12]:
            lines.append(
                f"- {item['title']}：收稿 {item['received'] or '无'}；录用 {item['accepted'] or '无'}；刊发 {item['published'] or '无'}；审稿天数 {item['review_days']}"
            )
    else:
        lines.append("- 暂无可识别日期线索。")
    lines.extend(["", "## 文章样本摘要"])
    for item in article_summaries:
        lines.append(f"- {item['title']}：章节 {item['heading_count']}；材料 {item['materials']}；方法 {item['methods']}")
    lines.extend([
        "",
        "## 人工复核建议",
        "- 规则识别的章节、材料和方法标记必须结合 RAG/PDF 全文复核。",
        "- 日期线索不完整时，不得计算或宣称审稿周期。",
        "- 少量全文样本不得代表全刊稳定风格。",
    ])

    Path(args.output_json).write_text(json.dumps(stats, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    Path(args.output_md).write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(json.dumps(stats, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
