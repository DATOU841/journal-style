#!/usr/bin/env python3
"""Analyze reference co-citation and source ecology from structured references."""

from __future__ import annotations

import argparse
import csv
import itertools
import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Analyze reference network.")
    parser.add_argument("--input", required=True, help="CSV/JSON/XLSX reference records.")
    parser.add_argument("--output-json", default="journal-reference-network-statistics.json")
    parser.add_argument("--output-md", default="journal-reference-network-report.md")
    return parser.parse_args()


def load_records(path: Path) -> list[dict]:
    suffix = path.suffix.lower()
    if suffix == ".json":
        data = json.loads(path.read_text(encoding="utf-8"))
        return list(data if isinstance(data, list) else data.get("records", []))
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as fh:
            return list(csv.DictReader(fh))
    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except Exception as exc:  # pragma: no cover - dependency guard
            raise SystemExit("openpyxl is required for xlsx input") from exc
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        headers = [str(cell).strip() if cell is not None else "" for cell in next(rows)]
        return [{headers[i]: row[i] for i in range(min(len(headers), len(row)))} for row in rows]
    raise SystemExit(f"Unsupported input format: {path.suffix}")


def pick(row: dict, *keys: str) -> str:
    lowered = {str(k).strip().lower(): v for k, v in row.items()}
    for key in keys:
        value = lowered.get(key.lower())
        if value is not None and str(value).strip():
            return str(value).strip()
    return ""


def normalize_text(value: str) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def split_reference_blob(blob: str) -> list[str]:
    if not blob:
        return []
    parts = [item.strip() for item in re.split(r"\|\||\n", blob) if item.strip()]
    return parts


def normalize_reference_title(row: dict, raw_reference: str = "") -> str:
    explicit = pick(row, "ref_title", "reference_title", "被引题名", "题名")
    if explicit:
        return normalize_text(explicit)
    raw = normalize_text(raw_reference or pick(row, "reference", "参考文献"))
    if not raw:
        return ""
    raw = re.sub(r"^[\[(（【]?\d+[\])）】.、\s]*", "", raw)
    return raw[:80]


def normalize_reference_author(row: dict, raw_reference: str = "") -> str:
    explicit = pick(row, "ref_author", "reference_author", "被引作者", "作者")
    if explicit:
        return normalize_text(explicit).split()[0]
    raw = normalize_text(raw_reference or pick(row, "reference", "参考文献"))
    if not raw:
        return ""
    raw = re.sub(r"^[\[(（【]?\d+[\])）】.、\s]*", "", raw)
    match = re.match(r"([^,，.。:：]{1,20})", raw)
    return normalize_text(match.group(1)) if match else ""


def parse_year(value: str) -> int | None:
    match = re.search(r"(19|20)\d{2}", value or "")
    return int(match.group(0)) if match else None


def is_truthy(value: str) -> bool:
    return str(value).strip() in {"1", "true", "True", "yes", "是", "Y", "y"}


def edge_key(left: str, right: str) -> str:
    first, second = sorted([left, right])
    return f"{first} -- {second}"


def ratio(numerator: int | float, denominator: int | float) -> float:
    return round(float(numerator) / float(denominator), 4) if denominator else 0.0


def expand_references(records: list[dict]) -> list[dict]:
    expanded: list[dict] = []
    for index, row in enumerate(records, start=1):
        article_id = pick(row, "article_id", "source_id", "题录ID") or f"row-{index}"
        blob = pick(row, "references", "参考文献列表")
        parts = split_reference_blob(blob)
        if parts:
            for part_index, part in enumerate(parts, start=1):
                item = dict(row)
                item["_article_id"] = article_id
                item["_reference_raw"] = part
                item["_reference_row_id"] = f"{article_id}#{part_index}"
                expanded.append(item)
        else:
            item = dict(row)
            item["_article_id"] = article_id
            item["_reference_raw"] = pick(row, "reference", "参考文献")
            item["_reference_row_id"] = f"{article_id}#1"
            expanded.append(item)
    return expanded


def evidence_strength(source_article_count: int, reference_count: int, title_missing_rate: float, year_missing_rate: float) -> str:
    if source_article_count >= 30 and reference_count >= 300 and title_missing_rate <= 0.2 and year_missing_rate <= 0.2:
        return "strong"
    if source_article_count >= 20 and reference_count >= 200 and title_missing_rate <= 0.3:
        return "medium"
    return "weak"


def main() -> int:
    args = parse_args()
    records = expand_references(load_records(Path(args.input)))

    article_refs: dict[str, set[str]] = defaultdict(set)
    article_authors: dict[str, set[str]] = defaultdict(set)
    ref_title_counts: Counter[str] = Counter()
    ref_author_counts: Counter[str] = Counter()
    ref_source_counts: Counter[str] = Counter()
    ref_type_counts: Counter[str] = Counter()
    language_counts: Counter[str] = Counter()
    year_counts: Counter[str] = Counter()
    co_citation_edges: Counter[str] = Counter()
    author_co_citation_edges: Counter[str] = Counter()
    self_journal_count = 0
    missing_title = 0
    missing_year = 0

    current_year = datetime.now().year
    recent_count = 0

    for row in records:
        article_id = str(row.get("_article_id", "")).strip() or "unknown"
        raw_reference = str(row.get("_reference_raw", "")).strip()
        title = normalize_reference_title(row, raw_reference)
        author = normalize_reference_author(row, raw_reference)
        year_value = pick(row, "ref_year", "reference_year", "被引年份", "year", "年份") or raw_reference
        year = parse_year(year_value)
        source = normalize_text(pick(row, "ref_source", "reference_source", "被引来源", "来源"))
        ref_type = normalize_text(pick(row, "ref_type", "reference_type", "文献类型"))
        language = normalize_text(pick(row, "language", "语言"))

        if not title:
            missing_title += 1
        else:
            ref_title_counts[title] += 1
            article_refs[article_id].add(title)

        if author:
            ref_author_counts[author] += 1
            article_authors[article_id].add(author)

        if source:
            ref_source_counts[source] += 1
        if ref_type:
            ref_type_counts[ref_type] += 1
        if language:
            language_counts[language] += 1

        if year is None:
            missing_year += 1
        else:
            year_counts[str(year)] += 1
            if year >= current_year - 4:
                recent_count += 1

        if is_truthy(pick(row, "is_self_journal", "self_journal_cite", "期刊内互引")):
            self_journal_count += 1

    for titles in article_refs.values():
        for left, right in itertools.combinations(sorted(titles), 2):
            co_citation_edges[edge_key(left, right)] += 1

    for authors in article_authors.values():
        for left, right in itertools.combinations(sorted(authors), 2):
            author_co_citation_edges[edge_key(left, right)] += 1

    source_article_count = len({str(row.get("_article_id", "")).strip() for row in records if row.get("_article_id")})
    title_missing_rate = ratio(missing_title, len(records))
    year_missing_rate = ratio(missing_year, len(records))
    strength = evidence_strength(source_article_count, len(records), title_missing_rate, year_missing_rate)
    core_candidates = []
    for title, count in ref_title_counts.most_common(30):
        source_span = sum(1 for refs in article_refs.values() if title in refs)
        if count >= 2 and source_span >= 2:
            core_candidates.append({"title": title, "count": count, "source_article_span": source_span})

    warnings = []
    if source_article_count < 20:
        warnings.append("来源文章低于 20 篇，只能写初步引用生态。")
    if len(records) < 200:
        warnings.append("参考文献记录低于 200 条，不得给强结论。")
    if title_missing_rate > 0.3:
        warnings.append("被引题名缺失率高于 30%，共引网络只能标待补解析。")
    if year_missing_rate > 0.3:
        warnings.append("被引年份缺失率高于 30%，年代结构只能标待补解析。")

    stats = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "source_article_count": source_article_count,
        "reference_record_count": len(records),
        "references_per_article_mean": round(len(records) / source_article_count, 2) if source_article_count else 0,
        "ref_title_missing_rate": title_missing_rate,
        "ref_year_missing_rate": year_missing_rate,
        "recent_reference_ratio": ratio(recent_count, len(records)),
        "self_journal_citation_ratio": ratio(self_journal_count, len(records)),
        "top_ref_authors": ref_author_counts.most_common(20),
        "top_ref_titles": ref_title_counts.most_common(20),
        "top_ref_sources": ref_source_counts.most_common(20),
        "ref_type_counts": dict(ref_type_counts.most_common()),
        "language_counts": dict(language_counts.most_common()),
        "year_counts": dict(year_counts.most_common()),
        "co_citation_edges": co_citation_edges.most_common(30),
        "author_co_citation_edges": author_co_citation_edges.most_common(30),
        "core_reference_candidates": core_candidates[:20],
        "evidence_strength": strength,
        "warnings": warnings,
    }

    lines = [
        "# 参考文献网络分析报告",
        "",
        f"- 来源文章数：{stats['source_article_count']}",
        f"- 参考文献记录数：{stats['reference_record_count']}",
        f"- 篇均参考文献数：{stats['references_per_article_mean']}",
        f"- 被引题名缺失率：{stats['ref_title_missing_rate']}",
        f"- 被引年份缺失率：{stats['ref_year_missing_rate']}",
        f"- 近五年引用比例：{stats['recent_reference_ratio']}",
        f"- 期刊内互引比例：{stats['self_journal_citation_ratio']}",
        f"- 证据强度：{stats['evidence_strength']}",
        "",
        "## 降级提示",
    ]
    lines.extend([f"- {item}" for item in warnings] or ["- 暂无阻断性降级提示。"])
    lines.append("")
    lines.append("## 高频被引作者")
    lines.extend([f"- {name}：{count}" for name, count in stats["top_ref_authors"][:10]])
    lines.append("")
    lines.append("## 高频被引题名")
    lines.extend([f"- {name}：{count}" for name, count in stats["top_ref_titles"][:10]])
    lines.append("")
    lines.append("## 高频被引来源")
    lines.extend([f"- {name}：{count}" for name, count in stats["top_ref_sources"][:10]] or ["- 来源字段不足。"])
    lines.append("")
    lines.append("## 共引边")
    lines.extend([f"- {edge}：{count}" for edge, count in stats["co_citation_edges"][:10]] or ["- 未形成可统计共引边。"])
    lines.append("")
    lines.append("## 核心文献候选（待核验）")
    lines.extend([f"- {item['title']}：{item['count']} 次，覆盖 {item['source_article_span']} 篇来源文章" for item in stats["core_reference_candidates"]] or ["- 样本不足，暂不列核心候选。"])

    Path(args.output_json).write_text(json.dumps(stats, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    Path(args.output_md).write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(json.dumps(stats, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
