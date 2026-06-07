#!/usr/bin/env python3
"""Analyze reference ecology from CSV/JSON/XLSX input."""

from __future__ import annotations

import argparse
import csv
import json
import statistics
from collections import Counter
from datetime import datetime
from pathlib import Path

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Analyze reference ecology.")
    parser.add_argument("--input", required=True, help="CSV/JSON/XLSX reference table.")
    parser.add_argument("--output-json", default="journal-reference-ecology-statistics.json")
    parser.add_argument("--output-md", default="journal-reference-ecology-report.md")
    return parser.parse_args()


def load_records(path: Path) -> list[dict]:
    suffix = path.suffix.lower()
    if suffix == ".json":
        data = json.loads(path.read_text(encoding="utf-8"))
        return list(data if isinstance(data, list) else data.get("records", []))
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as fh:
            return list(csv.DictReader(fh))
    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except Exception as exc:  # pragma: no cover - dependency guard
            raise SystemExit("openpyxl is required for xlsx input") from exc
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        headers = [str(cell).strip() if cell is not None else "" for cell in next(rows)]
        records = []
        for row in rows:
            record = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
            records.append(record)
        return records
    raise SystemExit(f"Unsupported input format: {path.suffix}")


def pick(row: dict, *keys: str) -> str:
    lowered = {str(k).strip().lower(): v for k, v in row.items()}
    for key in keys:
        if key.lower() in lowered and lowered[key.lower()] is not None:
            return str(lowered[key.lower()]).strip()
    return ""


def main() -> int:
    args = parse_args()
    records = load_records(Path(args.input))

    ref_counts = []
    year_counts = Counter()
    language_counts = Counter()
    author_counts = Counter()
    self_cite = 0

    current_year = datetime.now().year

    for row in records:
        ref_count_raw = pick(row, "reference_count", "参考文献数", "refs", "reference count")
        if ref_count_raw:
            try:
                ref_counts.append(float(ref_count_raw))
            except ValueError:
                pass

        year = pick(row, "year", "年份")
        if year:
            year_counts[year] += 1

        lang = pick(row, "language", "语言", "文献语言")
        if lang:
            language_counts[lang] += 1

        author = pick(row, "author", "作者", "第一作者")
        if author:
            author_counts[author] += 1

        if pick(row, "self_journal_cite", "期刊内互引", "is_self_journal") in {"1", "true", "True", "yes", "是"}:
            self_cite += 1

    recent_rows = 0
    for year, count in year_counts.items():
        try:
            if int(float(year)) >= current_year - 4:
                recent_rows += count
        except ValueError:
            continue

    stats = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "record_count": len(records),
        "reference_count_mean": round(statistics.mean(ref_counts), 2) if ref_counts else None,
        "reference_count_median": round(statistics.median(ref_counts), 2) if ref_counts else None,
        "reference_count_min": min(ref_counts) if ref_counts else None,
        "reference_count_max": max(ref_counts) if ref_counts else None,
        "language_counts": dict(language_counts.most_common()),
        "year_counts": dict(year_counts.most_common()),
        "recent_ratio": round(recent_rows / len(records), 4) if records else 0,
        "self_journal_citation_count": self_cite,
        "top_authors": author_counts.most_common(20),
    }

    report_lines = [
        "# 参考文献生态报告",
        "",
        f"- 样本数：{stats['record_count']}",
        f"- 参考文献均值：{stats['reference_count_mean']}",
        f"- 近五年比例：{stats['recent_ratio']}",
        f"- 期刊内互引数：{stats['self_journal_citation_count']}",
        "",
        "## 语言分布",
    ]
    report_lines.extend([f"- {k}：{v}" for k, v in stats["language_counts"].items()])
    report_lines.append("")
    report_lines.append("## 高频作者")
    report_lines.extend([f"- {k}：{v}" for k, v in stats["top_authors"][:10]])

    Path(args.output_json).write_text(json.dumps(stats, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    Path(args.output_md).write_text("\n".join(report_lines) + "\n", encoding="utf-8")
    print(json.dumps(stats, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
