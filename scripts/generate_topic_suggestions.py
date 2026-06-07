#!/usr/bin/env python3
"""Generate structured topic suggestions without drafting manuscript titles."""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter
from datetime import datetime
from pathlib import Path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate topic suggestions from corpus keywords.")
    parser.add_argument("--journal-profile", default="", help="Optional journal profile JSON.")
    parser.add_argument("--title-list", required=True, help="CSV/JSON/XLSX title list.")
    parser.add_argument("--user-keywords", required=True, help="Comma-separated user keywords.")
    parser.add_argument("--output", required=True, help="Output Markdown path.")
    parser.add_argument("--output-json", help="Optional output JSON path.")
    return parser.parse_args()


def load_records(path: Path) -> list[dict]:
    suffix = path.suffix.lower()
    if suffix == ".json":
        data = json.loads(path.read_text(encoding="utf-8"))
        return list(data if isinstance(data, list) else data.get("records", []))
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as fh:
            return list(csv.DictReader(fh))
    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except Exception as exc:
            raise SystemExit("openpyxl is required for xlsx input") from exc
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        headers = [str(cell).strip() if cell is not None else "" for cell in next(rows)]
        return [{headers[i]: row[i] for i in range(min(len(headers), len(row)))} for row in rows]
    raise SystemExit(f"Unsupported input format: {path.suffix}")


def pick(row: dict, *keys: str) -> str:
    lowered = {str(k).strip().lower(): v for k, v in row.items()}
    for key in keys:
        value = lowered.get(key.lower())
        if value is not None and str(value).strip():
            return str(value).strip()
    return ""


def classify(count_recent: int, count_total: int, years_seen: int) -> str:
    if count_recent >= 10:
        return "拥挤"
    if years_seen >= 5 and 5 <= count_total <= 15:
        return "稳定"
    if count_total < 3:
        return "边缘"
    return "可补强"


def parse_year(text: str) -> int | None:
    try:
        return int(float(text))
    except (TypeError, ValueError):
        return None


def main() -> int:
    args = parse_args()
    records = load_records(Path(args.title_list))
    user_keywords = [item.strip() for item in re.split(r"[,，、;；\s]+", args.user_keywords) if item.strip()]
    current_year = max([year for row in records if (year := parse_year(pick(row, "year", "年份"))) is not None] or [datetime.now().year])

    suggestions = []
    for keyword in user_keywords:
        matches = []
        years = set()
        recent = 0
        total_matches = 0
        for row in records:
            text = " ".join([pick(row, "title", "题名"), pick(row, "keywords", "关键词"), pick(row, "abstract", "摘要")])
            if keyword and keyword in text:
                total_matches += 1
                title = pick(row, "title", "题名")
                year = parse_year(pick(row, "year", "年份"))
                if year is not None:
                    years.add(year)
                    if year >= current_year - 2:
                        recent += 1
                if title and len(matches) < 5:
                    matches.append(title)
        status = classify(recent, total_matches, len(years))
        suggestion_type = "高适配" if status == "稳定" else "不建议" if status in {"拥挤", "边缘"} else "可补强"
        suggestions.append({
            "keyword": keyword,
            "suggestion_type": suggestion_type,
            "crowding_status": status,
            "recent_match_count": recent,
            "total_match_count": total_matches,
            "sample_titles": matches,
            "title_structure": "对象 + 问题动作 + 材料",
            "evidence_strength": "medium" if matches else "weak",
            "next_search": [keyword, f"{keyword} 材料", f"{keyword} 研究史"],
        })

    lines = ["# 选题建议报告", ""]
    for item in suggestions:
        lines.extend([
            f"## {item['keyword']}",
            "",
            f"- 建议类型：{item['suggestion_type']}",
            f"- 拥挤度：{item['crowding_status']}（近三年命中 {item['recent_match_count']} 条）",
            f"- 题名结构建议：{item['title_structure']}",
            f"- 证据强度：{item['evidence_strength']}",
            "- 同类题名样本：",
        ])
        lines.extend([f"  - {title}" for title in item["sample_titles"]] or ["  - 暂无直接样本，需补检。"])
        lines.append("- 补检方向：" + "；".join(item["next_search"]))
        lines.append("")

    Path(args.output).write_text("\n".join(lines) + "\n", encoding="utf-8")
    if args.output_json:
        Path(args.output_json).write_text(json.dumps({"suggestions": suggestions}, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({"suggestions": suggestions}, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
