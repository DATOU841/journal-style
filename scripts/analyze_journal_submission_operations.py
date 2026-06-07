#!/usr/bin/env python3
"""Analyze journal submission operations and public reputation evidence."""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter
from datetime import datetime
from pathlib import Path


SOURCE_LEVELS = {
    "official": 1,
    "database": 2,
    "third_party": 3,
    "forum": 4,
    "social": 4,
    "unknown": 5,
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Analyze journal submission operations evidence.")
    parser.add_argument("--input", required=True, help="CSV/JSON/XLSX evidence records.")
    parser.add_argument("--output-json", default="journal-submission-operations-statistics.json")
    parser.add_argument("--output-md", default="journal-submission-operations-report.md")
    return parser.parse_args()


def load_records(path: Path) -> list[dict]:
    suffix = path.suffix.lower()
    if suffix == ".json":
        data = json.loads(path.read_text(encoding="utf-8"))
        return list(data if isinstance(data, list) else data.get("records", []))
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as fh:
            return list(csv.DictReader(fh))
    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except Exception as exc:  # pragma: no cover - dependency guard
            raise SystemExit("openpyxl is required for xlsx input") from exc
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        headers = [str(cell).strip() if cell is not None else "" for cell in next(rows)]
        return [{headers[i]: row[i] for i in range(min(len(headers), len(row)))} for row in rows]
    raise SystemExit(f"Unsupported input format: {path.suffix}")


def pick(row: dict, *keys: str) -> str:
    lowered = {str(k).strip().lower(): v for k, v in row.items()}
    for key in keys:
        value = lowered.get(key.lower())
        if value is not None and str(value).strip():
            return str(value).strip()
    return ""


def normalize_source_type(value: str) -> str:
    text = (value or "unknown").strip().lower()
    if text in {"official", "官网", "投稿系统", "author_guidelines"}:
        return "official"
    if text in {"database", "数据库", "doi", "platform"}:
        return "database"
    if text in {"third_party", "第三方", "scirev", "publisher_metric"}:
        return "third_party"
    if text in {"forum", "论坛", "experience"}:
        return "forum"
    if text in {"social", "公众号", "社交媒体"}:
        return "social"
    return text if text in SOURCE_LEVELS else "unknown"


def parse_float(value: str) -> float | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    return float(match.group(0)) if match else None


def parse_bool(value: str) -> bool:
    return str(value).strip().lower() in {"1", "true", "yes", "y", "是", "有", "warning", "risk"}


def median(values: list[float]) -> float | None:
    if not values:
        return None
    ordered = sorted(values)
    middle = len(ordered) // 2
    if len(ordered) % 2:
        return round(ordered[middle], 2)
    return round((ordered[middle - 1] + ordered[middle]) / 2, 2)


def mean(values: list[float]) -> float | None:
    return round(sum(values) / len(values), 2) if values else None


def metric_summary(records: list[dict], key: str) -> dict:
    by_source: dict[str, list[float]] = {name: [] for name in SOURCE_LEVELS}
    for row in records:
        source_type = normalize_source_type(pick(row, "source_type", "来源类型"))
        value = parse_float(pick(row, key, key.replace("_", " "), key.upper()))
        if value is not None:
            by_source.setdefault(source_type, []).append(value)
    return {
        source: {
            "count": len(values),
            "mean": mean(values),
            "median": median(values),
            "min": min(values) if values else None,
            "max": max(values) if values else None,
        }
        for source, values in by_source.items()
        if values
    }


def evidence_strength(source_counts: Counter[str], has_official_timing: bool, has_official_fee_or_system: bool) -> str:
    if source_counts.get("official", 0) >= 1 and (has_official_timing or has_official_fee_or_system):
        return "strong"
    if source_counts.get("database", 0) + source_counts.get("third_party", 0) >= 2:
        return "medium"
    return "weak"


def main() -> int:
    args = parse_args()
    records = load_records(Path(args.input))

    source_counts: Counter[str] = Counter()
    reputation_counts: Counter[str] = Counter()
    fee_policies: Counter[str] = Counter()
    submission_systems: Counter[str] = Counter()
    warning_records: list[dict] = []
    reputation_notes: list[dict] = []
    official_urls: list[str] = []
    third_party_urls: list[str] = []

    has_official_timing = False
    has_official_fee_or_system = False

    for row in records:
        source_type = normalize_source_type(pick(row, "source_type", "来源类型"))
        source_counts[source_type] += 1
        source_url = pick(row, "source_url", "url", "链接")
        if source_url and source_type == "official":
            official_urls.append(source_url)
        elif source_url and source_type in {"third_party", "forum", "social"}:
            third_party_urls.append(source_url)

        timing_values = [
            parse_float(pick(row, "review_cycle_days")),
            parse_float(pick(row, "first_decision_days")),
            parse_float(pick(row, "acceptance_days")),
            parse_float(pick(row, "publication_lag_days")),
        ]
        if source_type == "official" and any(value is not None for value in timing_values):
            has_official_timing = True

        fee_policy = pick(row, "fee_policy", "收费政策")
        page_charge = pick(row, "page_charge", "版面费")
        apc = pick(row, "apc", "APC")
        submission_system = pick(row, "submission_system", "投稿系统")
        if fee_policy:
            fee_policies[fee_policy] += 1
        if page_charge:
            fee_policies[f"版面费：{page_charge}"] += 1
        if apc:
            fee_policies[f"APC：{apc}"] += 1
        if submission_system:
            submission_systems[submission_system] += 1
        if source_type == "official" and (fee_policy or page_charge or apc or submission_system):
            has_official_fee_or_system = True

        polarity = pick(row, "reputation_polarity", "声誉倾向").lower() or "neutral"
        if polarity not in {"positive", "negative", "mixed", "neutral"}:
            polarity = "neutral"
        reputation_counts[polarity] += 1
        note = pick(row, "reputation_note", "声誉说明", "note", "备注")
        if note and len(reputation_notes) < 20:
            reputation_notes.append({
                "source_type": source_type,
                "polarity": polarity,
                "note": note,
                "url": source_url,
            })

        warning = pick(row, "warning_flag", "风险标记")
        if parse_bool(warning):
            warning_records.append({
                "source_type": source_type,
                "title": pick(row, "source_title", "来源标题"),
                "note": note or warning,
                "url": source_url,
            })

    timing_metrics = {
        "review_cycle_days": metric_summary(records, "review_cycle_days"),
        "first_decision_days": metric_summary(records, "first_decision_days"),
        "acceptance_days": metric_summary(records, "acceptance_days"),
        "publication_lag_days": metric_summary(records, "publication_lag_days"),
    }
    strength = evidence_strength(source_counts, has_official_timing, has_official_fee_or_system)
    warnings = []
    if not source_counts.get("official"):
        warnings.append("缺少官网或投稿系统证据，投稿运营结论不得写成强结论。")
    if source_counts.get("forum", 0) or source_counts.get("social", 0):
        warnings.append("论坛、公众号或社交媒体内容只能作为待核验风评线索。")
    if reputation_counts.get("negative", 0) and not source_counts.get("official"):
        warnings.append("存在负面风评但缺少官方核验，应列为待核验风险，不得写成定性指控。")
    if not has_official_timing:
        warnings.append("缺少官方审稿周期证据，周期数据应分列为第三方或作者经验线索。")

    stats = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "record_count": len(records),
        "source_counts": dict(source_counts),
        "official_urls": list(dict.fromkeys(official_urls))[:10],
        "third_party_urls": list(dict.fromkeys(third_party_urls))[:10],
        "timing_metrics": timing_metrics,
        "fee_policy_counts": dict(fee_policies.most_common()),
        "submission_system_counts": dict(submission_systems.most_common()),
        "reputation_counts": dict(reputation_counts),
        "reputation_notes": reputation_notes,
        "warning_records": warning_records[:20],
        "evidence_strength": strength,
        "warnings": warnings,
    }

    lines = [
        "# 期刊投稿运营与公开声誉证据报告",
        "",
        f"- 证据记录数：{stats['record_count']}",
        f"- 来源构成：{stats['source_counts']}",
        f"- 证据强度：{stats['evidence_strength']}",
        "",
        "## 降级提示",
    ]
    lines.extend([f"- {item}" for item in warnings] or ["- 暂无阻断性降级提示。"])
    lines.extend(["", "## 官方入口"])
    lines.extend([f"- {url}" for url in stats["official_urls"]] or ["- 暂无官方入口证据。"])
    lines.extend(["", "## 审稿与见刊周期"])
    for metric_name, by_source in timing_metrics.items():
        lines.append(f"### {metric_name}")
        if not by_source:
            lines.append("- 暂无可统计数据。")
            continue
        for source_type, item in by_source.items():
            lines.append(
                f"- {source_type}：样本 {item['count']}，均值 {item['mean']} 天，中位数 {item['median']} 天，范围 {item['min']}-{item['max']} 天"
            )
    lines.extend(["", "## 费用政策"])
    lines.extend([f"- {name}：{count}" for name, count in fee_policies.most_common()] or ["- 暂无费用证据。"])
    lines.extend(["", "## 投稿系统"])
    lines.extend([f"- {name}：{count}" for name, count in submission_systems.most_common()] or ["- 暂无投稿系统证据。"])
    lines.extend(["", "## 公开声誉线索"])
    lines.append(f"- 声誉倾向计数：{stats['reputation_counts']}")
    for item in reputation_notes[:10]:
        lines.append(f"- [{item['source_type']}/{item['polarity']}] {item['note']} {item['url']}".strip())
    lines.extend(["", "## 待核验风险"])
    lines.extend([f"- [{item['source_type']}] {item['note']} {item['url']}".strip() for item in warning_records] or ["- 暂无风险标记。"])
    lines.extend([
        "",
        "## 写法边界",
        "- 第三方平台、论坛、公众号或社交媒体信息只能作为经验线索。",
        "- 负面风评不得写成定性指控；必须保留来源、时间和待核验状态。",
        "- 审稿周期和费用政策优先以官网、投稿系统或正式投稿须知为准。",
    ])

    Path(args.output_json).write_text(json.dumps(stats, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    Path(args.output_md).write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(json.dumps(stats, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
