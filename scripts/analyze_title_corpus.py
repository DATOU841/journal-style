#!/usr/bin/env python3
"""Analyze journal title corpus statistics from CSV/JSON/XLSX input."""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter
from datetime import datetime
from pathlib import Path

TITLE_SUFFIXES = [
    "研究", "探析", "探讨", "考论", "考察", "阐释", "辨析", "述评", "刍议",
    "论", "重审", "生成", "传播", "接受", "路径", "机制", "视域", "回顾",
]

MATERIAL_WORDS = ["档案", "文献", "碑帖", "题跋", "图像", "地方志", "手稿", "日记", "书信", "拓本", "印谱", "个案", "材料"]
ACTION_WORDS = ["考论", "考察", "辨析", "阐释", "重审", "探析", "探讨", "研究", "接受", "传播", "生成", "流变", "转化"]
TIME_WORDS = ["先秦", "汉代", "唐代", "宋代", "元代", "明代", "清代", "晚清", "民国", "近代", "现代", "当代"]
REGION_WORDS = ["中国", "岭南", "江南", "山东", "北京", "上海", "广东", "港澳", "东亚", "日本", "地方", "区域"]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Analyze title corpus.")
    parser.add_argument("--input", required=True, help="CSV/JSON/XLSX title corpus.")
    parser.add_argument("--output-json", default="journal-title-style-statistics.json")
    parser.add_argument("--output-md", default="journal-title-style-report.md")
    return parser.parse_args()


def load_records(path: Path) -> list[dict]:
    suffix = path.suffix.lower()
    if suffix == ".json":
        data = json.loads(path.read_text(encoding="utf-8"))
        return list(data if isinstance(data, list) else data.get("records", []))
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as fh:
            return list(csv.DictReader(fh))
    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except Exception as exc:  # pragma: no cover - dependency guard
            raise SystemExit("openpyxl is required for xlsx input") from exc
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        headers = [str(cell).strip() if cell is not None else "" for cell in next(rows)]
        records = []
        for row in rows:
            record = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
            records.append(record)
        return records
    raise SystemExit(f"Unsupported input format: {path.suffix}")


def pick(row: dict, *keys: str) -> str:
    lowered = {str(k).strip().lower(): v for k, v in row.items()}
    for key in keys:
        if key.lower() in lowered and lowered[key.lower()] is not None:
            return str(lowered[key.lower()]).strip()
    return ""


def classify_title_pattern(title: str) -> str:
    clean = re.sub(r"\s+", "", title)
    has_material = any(word in clean for word in MATERIAL_WORDS)
    has_action = any(word in clean for word in ACTION_WORDS)
    has_time = any(word in clean for word in TIME_WORDS)
    has_region = any(word in clean for word in REGION_WORDS)
    has_concept = bool(re.search(r"[“\"《][^”\"》]+[”\"》]", clean)) or any(word in clean for word in ["概念", "范畴", "观念", "理论", "话语"])
    if has_material and has_action:
        return "对象+问题动作+材料"
    if has_concept and (has_time or has_region):
        return "概念+时代/地区+方法"
    if has_time and has_region:
        return "时代+地区+对象"
    if "以" in clean and any(mark in clean for mark in ["为例", "为中心", "为对象"]):
        return "以...为例/中心/对象"
    if any(mark in clean for mark in ["：", ":"]):
        return "主副标题"
    return "其他"


def main() -> int:
    args = parse_args()
    path = Path(args.input)
    records = load_records(path)

    titles = [pick(r, "title", "题名") for r in records if pick(r, "title", "题名")]
    years = [pick(r, "year", "年份") for r in records if pick(r, "year", "年份")]
    columns = [pick(r, "column", "栏目") for r in records if pick(r, "column", "栏目")]
    keywords = [pick(r, "keywords", "关键词") for r in records if pick(r, "keywords", "关键词")]

    lengths = [len(re.sub(r"\s+", "", re.sub(r"[：:—\-]", "", t))) for t in titles]
    subtitle_count = sum(1 for t in titles if any(mark in t for mark in ["：", ":", "—", "-"]))

    suffix_counter = Counter()
    verb_counter = Counter()
    pattern_counter = Counter()
    punctuation_counter = Counter()
    for title in titles:
        clean = re.sub(r"\s+", "", title)
        pattern_counter[classify_title_pattern(clean)] += 1
        if "：" in title or ":" in title:
            punctuation_counter["冒号"] += 1
        if "—" in title or "-" in title:
            punctuation_counter["破折号"] += 1
        for suffix in TITLE_SUFFIXES:
            if clean.endswith(suffix) or suffix in clean:
                suffix_counter[suffix] += 1
        for verb in ["研究", "探析", "探讨", "考论", "考察", "阐释", "辨析", "述评", "刍议", "论"]:
            if verb in clean:
                verb_counter[verb] += 1

    year_counter = Counter(y for y in years if y)
    column_counter = Counter(c for c in columns if c)
    keyword_counter = Counter()
    for item in keywords:
        parts = [p.strip() for p in re.split(r"[、,，;；\s]+", item) if p.strip()]
        keyword_counter.update(parts)

    stats = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "record_count": len(records),
        "title_count": len(titles),
        "avg_title_length": round(sum(lengths) / len(lengths), 2) if lengths else 0,
        "median_title_length": sorted(lengths)[len(lengths) // 2] if lengths else 0,
        "subtitle_ratio": round(subtitle_count / len(titles), 4) if titles else 0,
        "top_suffixes": suffix_counter.most_common(15),
        "top_verbs": verb_counter.most_common(15),
        "title_structure_patterns": pattern_counter.most_common(),
        "punctuation_patterns": punctuation_counter.most_common(),
        "year_counts": dict(year_counter.most_common()),
        "column_counts": dict(column_counter.most_common()),
        "keyword_counts": keyword_counter.most_common(30),
    }

    report_lines = [
        "# 题名风格与基础趋势报告",
        "",
        f"- 样本数：{stats['title_count']}",
        f"- 平均题名长度：{stats['avg_title_length']}",
        f"- 主副标题比例：{stats['subtitle_ratio']}",
        "",
        "## 高频动词",
    ]
    report_lines.extend([f"- {k}：{v}" for k, v in stats["top_verbs"][:10]])
    report_lines.append("")
    report_lines.append("## 高频后缀 / 结构")
    report_lines.extend([f"- {k}：{v}" for k, v in stats["top_suffixes"][:10]])
    report_lines.append("")
    report_lines.append("## 题名结构模式")
    report_lines.extend([f"- {k}：{v}" for k, v in stats["title_structure_patterns"]])
    report_lines.append("")
    report_lines.append("## 标点结构")
    report_lines.extend([f"- {k}：{v}" for k, v in stats["punctuation_patterns"]] or ["- 暂无明显冒号或破折号结构。"])
    report_lines.append("")
    report_lines.append("## 年度分布")
    report_lines.extend([f"- {k}：{v}" for k, v in stats["year_counts"].items()])

    Path(args.output_json).write_text(json.dumps(stats, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    Path(args.output_md).write_text("\n".join(report_lines) + "\n", encoding="utf-8")
    print(json.dumps(stats, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
