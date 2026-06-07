#!/usr/bin/env python3
"""Analyze journal column lifecycle and stability."""

from __future__ import annotations

import argparse
import csv
import json
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Analyze journal column structure.")
    parser.add_argument("--title-list", required=True, help="CSV/JSON/XLSX title list.")
    parser.add_argument("--user-keywords", default="", help="Optional comma-separated user keywords.")
    parser.add_argument("--output-md", required=True, help="Output Markdown path.")
    parser.add_argument("--output-json", required=True, help="Output JSON path.")
    return parser.parse_args()


def load_records(path: Path) -> list[dict]:
    suffix = path.suffix.lower()
    if suffix == ".json":
        data = json.loads(path.read_text(encoding="utf-8"))
        return list(data if isinstance(data, list) else data.get("records", []))
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as fh:
            return list(csv.DictReader(fh))
    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except Exception as exc:
            raise SystemExit("openpyxl is required for xlsx input") from exc
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        headers = [str(cell).strip() if cell is not None else "" for cell in next(rows)]
        return [{headers[i]: row[i] for i in range(min(len(headers), len(row)))} for row in rows]
    raise SystemExit(f"Unsupported input format: {path.suffix}")


def pick(row: dict, *keys: str) -> str:
    lowered = {str(k).strip().lower(): v for k, v in row.items()}
    for key in keys:
        value = lowered.get(key.lower())
        if value is not None and str(value).strip():
            return str(value).strip()
    return ""


def main() -> int:
    args = parse_args()
    records = load_records(Path(args.title_list))
    by_column: dict[str, list[int]] = defaultdict(list)
    column_titles: dict[str, list[str]] = defaultdict(list)
    all_years = set()
    for row in records:
        column = pick(row, "column", "栏目") or "未标栏目"
        year_text = pick(row, "year", "年份")
        try:
            year = int(float(year_text))
        except ValueError:
            continue
        by_column[column].append(year)
        all_years.add(year)
        title = pick(row, "title", "题名")
        if title and len(column_titles[column]) < 5:
            column_titles[column].append(title)

    total_years = len(all_years) or 1
    stats = {}
    for column, years in by_column.items():
        year_counter = Counter(years)
        first_year = min(years)
        last_year = max(years)
        coverage = len(year_counter) / total_years
        lifecycle = last_year - first_year + 1
        kind = "核心栏目" if coverage >= 0.8 else "稳定栏目" if lifecycle >= 3 and len(year_counter) >= 3 else "临时栏目"
        stats[column] = {
            "first_year": first_year,
            "last_year": last_year,
            "year_count": len(year_counter),
            "article_count": len(years),
            "coverage_rate": round(coverage, 4),
            "lifecycle_years": lifecycle,
            "column_type": kind,
            "sample_titles": column_titles[column],
        }

    user_keywords = [item.strip() for item in args.user_keywords.split(",") if item.strip()]
    keyword_matches = []
    for column, titles in column_titles.items():
        score = sum(1 for title in titles for keyword in user_keywords if keyword in title)
        if score:
            keyword_matches.append({"column": column, "score": score})

    result = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "year_count": total_years,
        "column_count": len(stats),
        "columns": stats,
        "user_keyword_column_matches": sorted(keyword_matches, key=lambda item: item["score"], reverse=True),
    }

    lines = ["# 栏目结构与生命周期报告", "", f"- 栏目数：{len(stats)}", f"- 覆盖年份数：{total_years}", ""]
    for column, item in stats.items():
        lines.extend([
            f"## {column}",
            "",
            f"- 类型：{item['column_type']}",
            f"- 首次出现年份：{item['first_year']}",
            f"- 最后出现年份：{item['last_year']}",
            f"- 年度覆盖率：{item['coverage_rate']}",
            f"- 发文数：{item['article_count']}",
            "",
        ])

    Path(args.output_json).write_text(json.dumps(result, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    Path(args.output_md).write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

